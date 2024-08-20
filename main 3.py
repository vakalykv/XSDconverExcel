import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Пространства имен XML Schema
namespaces = {'xs': 'http://www.w3.org/2001/XMLSchema'}

# Функция для парсинга XML схемы
def parse_xml_schema(file_path):
    # Парсинг XML схемы
    tree = ET.parse(file_path)
    root = tree.getroot()

    elements = []

    # Рекурсивная функция для извлечения информации об элементах
    def extract_elements(element, xpath=""):
        # Для каждого элемента (xs:element)

        name = element.attrib.get('name')
        element_type = element.get('type')
        min_occurs = element.get('minOccurs', '1')
        max_occurs = element.get('maxOccurs', '1')

            # Обязательный элемент
        mandatory = 'О' if min_occurs == '1' else 'Н'

            # Признак повторяемости
        repeatability = 'Да' if max_occurs != '1' else 'Нет'

            # Аннотации
            # annotations = {'MDR': '', 'NSDR': ''}
            # annotation_elem = child.find('xs:annotation/xs:documentation', namespaces)
            # if annotation_elem is not None:
            #     lang = annotation_elem.get('{http://www.w3.org/XML/1998/namespace}lang')
            #     if lang == 'RusEng':
            #         annotations['MDR'] = annotation_elem.text
            #     elif lang == 'Rus':
            #         annotations['NSDR'] = annotation_elem.text

            # Xpath
        xpath += f"/{name}" if name else "" #= f"{xpath}/{name}" if name else ""
        print(f"xpath {xpath} внутри цикла.")
            # Извлечение возможных значений для перечислений (enumeration)
            #enumeration_values = []
            #simple_type = child.find('xs:simpleType', namespaces)
            #if simple_type is not None:
            #   for enum in simple_type.findall('xs:restriction/xs:enumeration', namespaces):
            #        enum_value = enum.get('value')
            #       enum_doc = enum.find('xs:annotation/xs:documentation', namespaces)
            #       enumeration_values.append(f"{enum_value} ({enum_doc.text})" if enum_doc is not None else enum_value)

            # Добавляем элемент в список
        elements.append({
                'name': name,
                'description': name, #annotation_elem.text if annotation_elem is not None else '',
                'mandatory': mandatory,
                'repeatability': repeatability,
                'xpath': xpath,
                'MDR': name, #annotations.get('MDR', ''),
                'NSDR': name, #annotations.get('NSDR', ''),
                'type': element_type, # if element_type else 'complex' if simple_type is None else 'simple',
                #'enumeration': name, #numeration_values,
        })
            # Обработка вложенных элементов в complexType (xs:complexType)
            # complex_type = child.find('xs:complexType', namespaces)
            # if complex_type is not None:
            #     sequence = complex_type.find('xs:sequence', namespaces)
            #     if sequence is not None:
            #         extract_elements(sequence, element_xpath)
            #     choice = complex_type.find('xs:choice', namespaces)
            #     if choice is not None:

        for child in element.findall(".//{<http://www.w3.org/2001/XMLSchema>}element"):
            extract_elements(child, xpath)

        # Запуск извлечения элементов начиная с корневого элемента
    extract_elements(root)
    return elements


# Функция для создания Excel файла
def create_excel(elements, output_file):
    # Создаем новый Excel файл
    workbook = Workbook()
    sheet = workbook.active

    # Заголовки столбцов
    headers = ["Номер по порядку", "Название элемента", "Описание элемента", "Признак обязательности", "Повторяемость",
               "XPath", "Пояснение MDR", "Пояснение НРД", "Тип данных"]

    sheet.append(headers)

    # Заполнение таблицы
    for idx, element in enumerate(elements, start=1):
        row = [
            idx,  # Номер по порядку
            element['name'],  # Название элемента
            element['description'],  # Описание элемента
            element['mandatory'],  # Признак обязательности
            element['repeatability'],  # Повторяемость
            element['xpath'],  # XPath
            element['MDR'],  # Пояснение MDR
            element['NSDR'],  # Пояснение НРД
            element['type'],  # Тип данных
            #', '.join(element['enumeration'])  # Enumeration values
        ]
        sheet.append(row)

    # Автоматическая настройка ширины столбцов
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Получаем букву колонки
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Сохранение файла
    workbook.save(output_file)

# Основной код
if __name__ == "__main__":
    # Путь к XML схеме
    xml_schema_file = "camt.053.001.06.xsd"

    # Парсим XML схему
    elements = parse_xml_schema(xml_schema_file)

    # Создаем Excel файл
    output_excel_file = "nalogadmin.xlsx"
    create_excel(elements, output_excel_file)

    print(f"Файл {output_excel_file} успешно создан.")
